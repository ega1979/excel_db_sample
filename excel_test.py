from openpyxl import load_workbook

EXCEL_FILE = 'tickets.xlsx'

def search_and_mark_attendance(ticket_number):
    # Load the workbook and select the active sheet
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    found = False

    # Iterate through the rows to find the ticket number
    for row in ws.iter_rows(min_row=2): # Start from the second row to skip headers
        if str(row[0].value) == ticket_number:
            name = row[1].value
            company = row[2].value
            print(f"紹介結果: {name} ({company})")

            row[3].value = "○" # Mark attendance with a check mark
            found = True
            break
            
    if found:
        wb.save(EXCEL_FILE) # Save the changes to the workbook
        print("出席マークを保存しました。") 
    else:
        print("チケット番号が見つかりませんでした。")

if __name__ == "__main__":
    ticket = input("チケット番号を入力してください: ")
    search_and_mark_attendance(ticket.strip())